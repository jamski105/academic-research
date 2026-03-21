#!/usr/bin/env python3
"""Excel literature spreadsheet generator — v4 new module.

Generates a professionally formatted Excel workbook with:
  Sheet 1: Literaturübersicht (full paper list with 5D scores)
  Sheet 2: Cluster-Analyse (cluster statistics + chart)
  Sheet 3: Kapitel-Zuordnung (chapter assignment from outline)
  Sheet 4: Datenblatt (hidden raw data)

Usage:
  python excel.py --papers ranked.json --output literature.xlsx
  python excel.py --papers ranked.json --output literature.xlsx --context context.json
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from text_utils import load_json

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Color definitions
# ---------------------------------------------------------------------------

COLORS = {
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "kernliteratur": "C6EFCE",
    "ergaenzungsliteratur": "BDD7EE",
    "hintergrundliteratur": "E2EFDA",
    "methodenliteratur": "FFF2CC",
    "score_high": "C6EFCE",
    "score_mid": "FFEB9C",
    "score_low": "FFC7CE",
    "alt_row": "F2F2F2",
}

CLUSTER_FILLS = {
    "Kernliteratur": PatternFill(start_color=COLORS["kernliteratur"], fill_type="solid"),
    "Ergänzungsliteratur": PatternFill(start_color=COLORS["ergaenzungsliteratur"], fill_type="solid"),
    "Hintergrundliteratur": PatternFill(start_color=COLORS["hintergrundliteratur"], fill_type="solid"),
    "Methodenliteratur": PatternFill(start_color=COLORS["methodenliteratur"], fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color=COLORS["header_bg"], fill_type="solid")
HEADER_FONT = Font(bold=True, color=COLORS["header_fg"], size=11)
ALT_ROW_FILL = PatternFill(start_color=COLORS["alt_row"], fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _score_fill(score: float) -> PatternFill:
    """Conditional color for score cells."""
    if score >= 0.7:
        return PatternFill(start_color=COLORS["score_high"], fill_type="solid")
    if score >= 0.4:
        return PatternFill(start_color=COLORS["score_mid"], fill_type="solid")
    return PatternFill(start_color=COLORS["score_low"], fill_type="solid")


# ---------------------------------------------------------------------------
# Sheet 1: Literaturübersicht
# ---------------------------------------------------------------------------

OVERVIEW_HEADERS = [
    ("#", 5),
    ("Cluster", 18),
    ("Titel", 55),
    ("Autoren", 35),
    ("Jahr", 8),
    ("Venue", 25),
    ("DOI", 25),
    ("Score", 8),
    ("Relevanz", 9),
    ("Aktualität", 10),
    ("Qualität", 9),
    ("Autorität", 9),
    ("Zugang", 8),
    ("Zitationen", 10),
    ("Quelle", 15),
    ("PDF", 8),
    ("Kapitel", 15),
    ("Notizen", 40),
]


def _write_overview(ws: Worksheet, papers: list[dict[str, Any]], context: dict | None) -> None:
    """Write the main literature overview sheet."""
    ws.title = "Literaturübersicht"

    # Header row
    for col_idx, (header, width) in enumerate(OVERVIEW_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OVERVIEW_HEADERS))}1"

    # Data rows
    for row_idx, paper in enumerate(papers, 2):
        scores = paper.get("scores") or {}
        cluster = paper.get("cluster", "Hintergrundliteratur")
        authors = ", ".join(paper.get("authors") or [])

        # Determine PDF status
        pdf_status = "nein"
        if paper.get("oa_url") or paper.get("open_access_pdf"):
            pdf_status = "ja"
        elif paper.get("doi") and any(v in (paper.get("venue") or "").lower() for v in ["springer", "ieee", "elsevier"]):
            pdf_status = "HAN"

        # Determine chapter assignment
        chapter = ""
        if context and context.get("chapters"):
            # Simple keyword matching for chapter assignment
            title_lower = (paper.get("title") or "").lower()
            for ch in context.get("chapters", []):
                if any(kw.lower() in title_lower for kw in ch.get("keywords", [])):
                    chapter = ch.get("name", "")
                    break

        row_data = [
            row_idx - 1,
            cluster,
            paper.get("title") or "Untitled",
            authors[:100],
            paper.get("year"),
            paper.get("venue") or "",
            paper.get("doi") or "",
            scores.get("total", 0),
            scores.get("relevance", 0),
            scores.get("recency", 0),
            scores.get("quality", 0),
            scores.get("authority", 0),
            scores.get("accessibility", 0),
            paper.get("citations", 0),
            paper.get("source_module", ""),
            pdf_status,
            chapter,
            "",  # Notes (empty for user)
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=col_idx in (3, 4))

            # Cluster color
            if col_idx == 2:
                fill = CLUSTER_FILLS.get(cluster)
                if fill:
                    cell.fill = fill

            # Score conditional formatting
            if col_idx in (8, 9, 10, 11, 12, 13) and isinstance(value, (int, float)):
                cell.fill = _score_fill(value)
                cell.number_format = "0.00"

            # Alternating row color for non-colored cells
            if (row_idx % 2 == 0) and col_idx not in (2, 8, 9, 10, 11, 12, 13):
                cell.fill = ALT_ROW_FILL


# ---------------------------------------------------------------------------
# Sheet 2: Cluster-Analyse
# ---------------------------------------------------------------------------

def _write_cluster_analysis(wb: openpyxl.Workbook, papers: list[dict[str, Any]]) -> None:
    """Write cluster statistics and chart."""
    ws = wb.create_sheet("Cluster-Analyse")

    # Gather stats
    clusters = ["Kernliteratur", "Ergänzungsliteratur", "Hintergrundliteratur", "Methodenliteratur"]
    stats: dict[str, dict[str, Any]] = {}
    for cluster in clusters:
        cluster_papers = [p for p in papers if p.get("cluster") == cluster]
        count = len(cluster_papers)
        avg_score = sum((p.get("scores") or {}).get("total", 0) for p in cluster_papers) / max(1, count)
        avg_recency = sum((p.get("scores") or {}).get("recency", 0) for p in cluster_papers) / max(1, count)
        peer_reviewed = sum(1 for p in cluster_papers if (p.get("scores") or {}).get("authority", 0) >= 0.7)
        pr_pct = (peer_reviewed / count * 100) if count > 0 else 0
        stats[cluster] = {"count": count, "avg_score": avg_score, "avg_recency": avg_recency, "pr_pct": pr_pct}

    headers = ["Cluster", "Anzahl", "Ø Score", "Ø Aktualität", "Peer-Reviewed %", "Beschreibung"]
    descriptions = {
        "Kernliteratur": "Direkt zur Forschungsfrage, muss zitiert werden",
        "Ergänzungsliteratur": "Unterstützend, Gegenargumente, Vertiefung",
        "Hintergrundliteratur": "Grundlagen, Lehrbücher, Standards",
        "Methodenliteratur": "Methodik-Begründung",
    }

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, cluster in enumerate(clusters, 2):
        s = stats[cluster]
        ws.cell(row=row_idx, column=1, value=cluster).fill = CLUSTER_FILLS.get(cluster, PatternFill())
        ws.cell(row=row_idx, column=2, value=s["count"])
        ws.cell(row=row_idx, column=3, value=round(s["avg_score"], 2)).number_format = "0.00"
        ws.cell(row=row_idx, column=4, value=round(s["avg_recency"], 2)).number_format = "0.00"
        ws.cell(row=row_idx, column=5, value=round(s["pr_pct"], 1)).number_format = "0.0"
        ws.cell(row=row_idx, column=6, value=descriptions.get(cluster, ""))
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Column widths
    for col, w in enumerate([20, 10, 10, 12, 15, 50], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Papers pro Cluster"
    chart.y_axis.title = "Anzahl"
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "A8")


# ---------------------------------------------------------------------------
# Sheet 3: Kapitel-Zuordnung
# ---------------------------------------------------------------------------

def _write_chapter_assignment(wb: openpyxl.Workbook, papers: list[dict[str, Any]], context: dict | None) -> None:
    """Write chapter-to-paper assignment sheet."""
    ws = wb.create_sheet("Kapitel-Zuordnung")

    headers = ["Kapitel", "Quellen", "Kernlit.", "Ergänz.", "Lücken", "Paper-Liste"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    if not context or not context.get("chapters"):
        ws.cell(row=2, column=1, value="Kein akademischer Kontext vorhanden. Bitte Academic Context Skill nutzen.")
        for col, w in enumerate([30, 10, 10, 10, 40, 60], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        return

    for row_idx, chapter in enumerate(context.get("chapters", []), 2):
        ch_name = chapter.get("name", "")
        keywords = chapter.get("keywords", [])
        matched = []
        for p in papers:
            title = (p.get("title") or "").lower()
            if any(kw.lower() in title for kw in keywords):
                matched.append(p)

        kern = sum(1 for p in matched if p.get("cluster") == "Kernliteratur")
        erg = sum(1 for p in matched if p.get("cluster") == "Ergänzungsliteratur")
        gaps = "Zu wenig Quellen" if len(matched) < 3 else ""
        paper_list = ", ".join(f"{(p.get('authors') or ['?'])[0].split()[-1]}{p.get('year', '')}" for p in matched[:10])

        ws.cell(row=row_idx, column=1, value=ch_name)
        ws.cell(row=row_idx, column=2, value=len(matched))
        ws.cell(row=row_idx, column=3, value=kern)
        ws.cell(row=row_idx, column=4, value=erg)
        ws.cell(row=row_idx, column=5, value=gaps)
        ws.cell(row=row_idx, column=6, value=paper_list)
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    for col, w in enumerate([30, 10, 10, 10, 40, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Sheet 4: Datenblatt (hidden)
# ---------------------------------------------------------------------------

def _write_raw_data(wb: openpyxl.Workbook, papers: list[dict[str, Any]]) -> None:
    """Write hidden raw data sheet."""
    ws = wb.create_sheet("Datenblatt")

    if not papers:
        return

    # Flatten papers to rows
    all_keys = set()
    for p in papers:
        all_keys.update(p.keys())
        if "scores" in p:
            for sk in (p.get("scores") or {}):
                all_keys.add(f"score_{sk}")
    all_keys.discard("scores")
    headers = sorted(all_keys)

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, paper in enumerate(papers, 2):
        flat = dict(paper)
        for sk, sv in (paper.get("scores") or {}).items():
            flat[f"score_{sk}"] = sv
        flat.pop("scores", None)
        flat["authors"] = ", ".join(flat.get("authors") or [])
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=flat.get(key))

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_excel(
    papers: list[dict[str, Any]],
    output_path: str,
    context: dict | None = None,
) -> str:
    """Generate the literature Excel workbook. Returns output path."""
    wb = openpyxl.Workbook()
    ws = wb.active

    _write_overview(ws, papers, context)
    _write_cluster_analysis(wb, papers)
    _write_chapter_assignment(wb, papers, context)
    _write_raw_data(wb, papers)

    wb.save(output_path)
    log.info("Excel saved: %s (%d papers)", output_path, len(papers))
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate literature Excel spreadsheet")
    parser.add_argument("--papers", required=True, help="Ranked papers JSON")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--context", help="Academic context JSON (for chapter assignment)")
    parser.add_argument("--scoring-config", help="Scoring config YAML (unused, for future)")
    return parser.parse_args()


def main() -> int:
    logging.basicConfig(level=logging.INFO)
    args = parse_args()

    try:
        papers = load_json(args.papers)
    except Exception:
        log.exception("Failed to load papers")
        return 1

    context = None
    if args.context:
        try:
            context = load_json(args.context)
        except Exception:
            log.warning("Failed to load context, continuing without")

    try:
        generate_excel(papers, args.output, context)
    except Exception:
        log.exception("Failed to generate Excel")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
